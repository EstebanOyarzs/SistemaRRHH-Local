"""
================================================================================
NORMALIZADOR DE CAPACITACION -> SQLITE
================================================================================
Cruza dos Excel de Dotacion (mes actual + mes anterior, "Detalle...") y arma
2 tablas (mas la dotacion completa):

    - capacitacion_dotacion        : dotacion del mes actual completa (sin clasificar)
    - capacitacion_nuevos_ingresos : personas cuya Fecha Ingreso cae en el mes del reporte
    - capacitacion_cambios_cargo   : personas presentes en ambas dotaciones cuya Funcion cambio

La Clasificacion de Procedimientos de cada persona y la lista de "cargos a
revisar" NO se guardan aca — se calculan en vivo en router.py haciendo JOIN
contra la tabla maestra capacitacion_procedimientos (editable desde la UI),
para que clasificar un cargo se refleje al instante sin tener que volver a
subir las dotaciones.

Cada corrida REEMPLAZA por completo esas 3 tablas (no la tabla maestra, que
se edita aparte). Antes de reemplazarlas se respalda la base anterior en
Capacitacion/data/backups/, igual que Sobretiempo.

Los dos Excel de dotacion cambian de nombre de hoja mes a mes ("Detalle
Activos", "Detalle activos", "Detalle Chilquintas y Filiales") y sus
encabezados se leen con acentos corruptos segun la codificacion de consola,
asi que las columnas se seleccionan por POSICION (estable en los 5 archivos
de 2026 verificados a mano), no por nombre.
"""

import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from sqlalchemy import Date, text

from backend.config import PROJECT_ROOT
from backend.dashboards.capacitacion.db import (
    DB_DIR,
    DB_PATH,
    TABLE_CAMBIOS_CARGO,
    TABLE_DOTACION,
    TABLE_NUEVOS_INGRESOS,
    TABLE_PROCEDIMIENTOS,
    engine,
)

# Posicion (0-indexada) de cada columna necesaria en la hoja "Detalle..." de
# los Excel de Dotacion. Verificado igual en Marzo/Abril/Mayo/Junio/Julio 2026.
COL_POSICIONES = {
    "Cod_Personal": 0,
    "Nombre_Completo": 1,
    "Area_Personal": 5,
    "Sociedad": 6,
    "Unidad_Organizativa": 10,
    "Funcion": 11,
    "Gerencia": 12,
    "Subgerencia": 13,
    "Fecha_Ingreso": 14,
    "Nombre_Superior": 32,
    "Nacionalidad": 33,
    "Mail": 34,
}
COLUMNAS_TEXTO = [
    "Nombre_Completo", "Area_Personal", "Sociedad", "Unidad_Organizativa",
    "Funcion", "Gerencia", "Subgerencia", "Nombre_Superior", "Nacionalidad", "Mail",
]

CLASIFICACION_DEFAULT = "Revisar si aplican procedimientos "

SEED_PATH = PROJECT_ROOT / "Capacitacion" / "Archivos Ejemplo" / "Reporte para capacitaciones - Mayo 2026.xlsx"

MESES_NOMBRE = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre",
    11: "Noviembre", 12: "Diciembre",
}


# ------------------------------------------------------------------------
# 1. LECTURA DE DOTACION
# ------------------------------------------------------------------------
def find_detalle_sheet(wb) -> str:
    for name in wb.sheetnames:
        if name.strip().lower().startswith("detalle"):
            return name
    raise ValueError(
        f"No se encontro una hoja 'Detalle...' en el archivo (hojas disponibles: {wb.sheetnames})"
    )


def cargar_dotacion(path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = find_detalle_sheet(wb)
    finally:
        wb.close()

    raw = pd.read_excel(path, sheet_name=sheet, header=0)
    cols = list(raw.columns)

    df = pd.DataFrame({
        nombre: raw[cols[pos]] for nombre, pos in COL_POSICIONES.items()
    })

    # Solo filas con Cod_Personal numerico (descarta filas vacias/basura)
    df = df[pd.to_numeric(df["Cod_Personal"], errors="coerce").notna()].copy()
    df["Cod_Personal"] = df["Cod_Personal"].astype("Int64")

    for col in COLUMNAS_TEXTO:
        df[col] = df[col].astype(str).str.strip()

    df["Fecha_Ingreso"] = pd.to_datetime(df["Fecha_Ingreso"], errors="coerce")

    df = df.drop_duplicates(subset="Cod_Personal", keep="last")
    return df.reset_index(drop=True)


# ------------------------------------------------------------------------
# 2. TABLA MAESTRA DE PROCEDIMIENTOS (Funcion -> Codigos)
# ------------------------------------------------------------------------
def sembrar_procedimientos_si_vacia() -> None:
    """Siembra capacitacion_procedimientos una sola vez, leyendo la hoja
    "Procedimientos" del Excel de referencia mas completo (Mayo 2026, 643
    funciones). Si la tabla ya tiene datos, o si el archivo de referencia no
    esta disponible, no hace nada — de ahi en mas la tabla se edita a mano
    desde la UI."""
    with engine.connect() as conn:
        count = conn.execute(text(f"SELECT COUNT(*) FROM {TABLE_PROCEDIMIENTOS}")).scalar()
    if count:
        return
    if not SEED_PATH.exists():
        return

    raw = pd.read_excel(SEED_PATH, sheet_name="Procedimientos", header=0, usecols=[0, 1])
    raw.columns = ["Funcion", "Codigos"]
    raw = raw.dropna(subset=["Funcion"])
    raw["Funcion"] = raw["Funcion"].astype(str).str.strip()
    raw["Codigos"] = raw["Codigos"].astype(str).str.strip()

    # Ante funciones duplicadas en la hoja de referencia, gana la ultima
    # aparicion (mismo criterio que un dict construido en orden de fila).
    mapa = dict(zip(raw["Funcion"], raw["Codigos"]))

    with engine.begin() as conn:
        for funcion, codigos in mapa.items():
            conn.execute(
                text(f"INSERT OR REPLACE INTO {TABLE_PROCEDIMIENTOS} (Funcion, Codigos) VALUES (:f, :c)"),
                {"f": funcion, "c": codigos},
            )


# ------------------------------------------------------------------------
# 3. NUEVOS INGRESOS + CAMBIOS DE CARGO
#    (la Clasificacion de Procedimientos y los "cargos a revisar" NO se
#    calculan aca — se hacen en vivo en router.py via JOIN contra la tabla
#    maestra, ver nota en db.py)
# ------------------------------------------------------------------------
def construir_nuevos_ingresos(df_dotacion: pd.DataFrame, mes_reporte: int, anio_reporte: int) -> pd.DataFrame:
    mask = (
        (df_dotacion["Fecha_Ingreso"].dt.month == mes_reporte)
        & (df_dotacion["Fecha_Ingreso"].dt.year == anio_reporte)
    )
    return df_dotacion[mask].reset_index(drop=True)


def construir_cambios_cargo(df_dotacion: pd.DataFrame, df_anterior: pd.DataFrame) -> pd.DataFrame:
    funcion_anterior = df_anterior.set_index("Cod_Personal")["Funcion"]
    df = df_dotacion.copy()
    df["Cargo_Anterior"] = df["Cod_Personal"].map(funcion_anterior)
    df = df[df["Cargo_Anterior"].notna() & (df["Cargo_Anterior"] != df["Funcion"])]
    return df.reset_index(drop=True)


# ------------------------------------------------------------------------
# 4. RESPALDO + CARGA A SQLITE
# ------------------------------------------------------------------------
BACKUPS_DIR = DB_DIR / "backups"


def respaldar_db() -> Path | None:
    if not DB_PATH.exists():
        return None
    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    destino = BACKUPS_DIR / f"capacitacion_{timestamp}.db"
    shutil.copy2(DB_PATH, destino)
    return destino


def guardar_sqlite(df_dotacion, df_nuevos, df_cambios) -> None:
    with engine.begin() as conn:
        df_dotacion.to_sql(
            TABLE_DOTACION, conn, if_exists="replace", index=False,
            dtype={"Fecha_Ingreso": Date()},
        )
        df_nuevos.to_sql(
            TABLE_NUEVOS_INGRESOS, conn, if_exists="replace", index=False,
            dtype={"Fecha_Ingreso": Date()},
        )
        df_cambios.to_sql(
            TABLE_CAMBIOS_CARGO, conn, if_exists="replace", index=False,
            dtype={"Fecha_Ingreso": Date()},
        )


def procesar_archivos(path_actual, path_anterior, mes_reporte: int, anio_reporte: int) -> dict:
    """Corre todo el pipeline y deja la base actualizada. Si el parsing de
    cualquiera de los dos archivos falla, tira la excepcion ANTES de tocar
    la base — el respaldo/reemplazo recien pasa al final."""
    df_actual = cargar_dotacion(path_actual)
    df_anterior = cargar_dotacion(path_anterior)

    df_actual["Mes_Reporte"] = mes_reporte
    df_actual["Anio_Reporte"] = anio_reporte

    df_nuevos = construir_nuevos_ingresos(df_actual, mes_reporte, anio_reporte)
    df_cambios = construir_cambios_cargo(df_actual, df_anterior)

    backup = respaldar_db()
    guardar_sqlite(df_actual, df_nuevos, df_cambios)

    return {
        "dotacion": len(df_actual),
        "nuevos_ingresos": len(df_nuevos),
        "cambios_cargo": len(df_cambios),
        "backup": str(backup) if backup else None,
    }
