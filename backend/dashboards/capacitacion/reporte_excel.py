"""
================================================================================
GENERADOR DE REPORTE EXCEL DE CAPACITACION
================================================================================
Arma el mismo Excel de 3 hojas (Resumen, Dotación CHTA, Procedimientos) que
"Capacitacion/Archivos Ejemplo/Reporte para capacitaciones - *.xlsx" — misma
logica que usa el endpoint GET /dashboards/capacitacion/exportar-excel
(backend/dashboards/capacitacion/router.py), centralizada aca para que tanto
el endpoint (descarga desde el navegador) como el script de linea de comandos
(Capacitacion/generar_reporte_capacitacion.py, sin backend corriendo) usen
exactamente la misma logica — nunca se duplica.
"""

from datetime import datetime
from pathlib import Path

import openpyxl
from sqlalchemy import text
from sqlalchemy.exc import OperationalError

from backend.config import PROJECT_ROOT
from backend.dashboards.capacitacion.db import (
    TABLE_CAMBIOS_CARGO,
    TABLE_DOTACION,
    TABLE_NUEVOS_INGRESOS,
    TABLE_PROCEDIMIENTOS,
    engine,
)
from backend.dashboards.capacitacion.normalizar import CLASIFICACION_DEFAULT, MESES_NOMBRE

REPORTES_DIR = PROJECT_ROOT / "Capacitacion" / "data" / "reportes"

COLUMNAS_DOTACION = [
    ("Cod_Personal", "N° pers."), ("Nombre_Completo", "Nombre completo"),
    ("Area_Personal", "Área de personal"), ("Sociedad", "Sociedad"),
    ("Unidad_Organizativa", "Unidad organizativa"), ("Funcion", "Función"),
    ("Gerencia", "Gerencia"), ("Subgerencia", "Subgerencia"),
    ("Fecha_Ingreso", "Fecha Ingreso"), ("Nombre_Superior", "Nombre del superior (GO)"),
    ("Nacionalidad", "Nacionalidad"), ("Mail", "Mail"),
    ("Clasificacion", "Clasificación de Procedimientos"),
]
COLUMNAS_CAMBIOS_CARGO = [
    ("Cod_Personal", "N° pers."), ("Nombre_Completo", "Nombre completo"),
    ("Area_Personal", "Área de personal"), ("Sociedad", "Sociedad"),
    ("Unidad_Organizativa", "Unidad organizativa"), ("Cargo_Anterior", "Cargo Anterior"),
    ("Funcion", "Función"), ("Gerencia", "Gerencia"), ("Subgerencia", "Subgerencia"),
    ("Fecha_Ingreso", "Fecha Ingreso"), ("Nombre_Superior", "Nombre del superior (GO)"),
    ("Nacionalidad", "Nacionalidad"), ("Mail", "Mail"),
    ("Clasificacion", "Clasificación de Procedimientos"),
]
COLUMNAS_PROCEDIMIENTOS = [("Funcion", "Función"), ("Codigos", "Procedimientos asignados")]


def _query_con_clasificacion(table: str) -> list[dict]:
    sql = (
        f"SELECT d.*, COALESCE(p.Codigos, :default) AS Clasificacion "
        f"FROM {table} d LEFT JOIN {TABLE_PROCEDIMIENTOS} p ON d.Funcion = p.Funcion"
    )
    try:
        with engine.connect() as conn:
            rows = conn.execute(text(sql), {"default": CLASIFICACION_DEFAULT}).mappings().all()
    except OperationalError:
        # Todavia no se cargo ninguna dotacion (la tabla ni existe) — se
        # trata igual que "sin datos", no como un error real.
        return []
    return [dict(row) for row in rows]


def obtener_datos_reporte() -> dict:
    """Trae dotacion/nuevos_ingresos/cambios_cargo (con Clasificacion
    calculada en vivo) + la tabla maestra de procedimientos. dotacion vacia
    ([]) significa que todavia no se cargo ninguna planilla."""
    dotacion = _query_con_clasificacion(TABLE_DOTACION)
    nuevos = _query_con_clasificacion(TABLE_NUEVOS_INGRESOS)
    cambios = _query_con_clasificacion(TABLE_CAMBIOS_CARGO)
    with engine.connect() as conn:
        procedimientos = [
            dict(row) for row in conn.execute(
                text(f"SELECT * FROM {TABLE_PROCEDIMIENTOS} ORDER BY Funcion")
            ).mappings().all()
        ]

    mes_reporte = dotacion[0].get("Mes_Reporte") if dotacion else None
    anio_reporte = dotacion[0].get("Anio_Reporte") if dotacion else None

    return {
        "dotacion": dotacion,
        "nuevos": nuevos,
        "cambios": cambios,
        "procedimientos": procedimientos,
        "mes_reporte": mes_reporte,
        "anio_reporte": anio_reporte,
    }


def nombre_archivo_reporte(mes_reporte, anio_reporte) -> str:
    mes_nombre = MESES_NOMBRE.get(mes_reporte, "")
    return f"Reporte para capacitaciones - {mes_nombre} {anio_reporte}.xlsx"


def _valor_celda(campo: str, valor):
    if campo == "Fecha_Ingreso" and isinstance(valor, str) and valor:
        try:
            return datetime.strptime(valor[:10], "%Y-%m-%d").date()
        except ValueError:
            return valor
    return valor


def _escribir_tabla(ws, start_row: int, columnas, filas: list[dict]) -> int:
    for j, (_, header) in enumerate(columnas, start=1):
        ws.cell(row=start_row, column=j, value=header)
    r = start_row + 1
    for fila in filas:
        for j, (campo, _) in enumerate(columnas, start=1):
            ws.cell(row=r, column=j, value=_valor_celda(campo, fila.get(campo)))
        r += 1
    return r


def construir_workbook(dotacion, nuevos, cambios, procedimientos) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    fila = 2
    ws_resumen.cell(row=fila, column=1, value="Nuevos ingresos ")
    fila = _escribir_tabla(ws_resumen, fila + 1, COLUMNAS_DOTACION, nuevos)
    fila += 2
    ws_resumen.cell(row=fila, column=1, value="Cambios de cargo")
    _escribir_tabla(ws_resumen, fila + 1, COLUMNAS_CAMBIOS_CARGO, cambios)

    ws_dotacion = wb.create_sheet("Dotación CHTA")
    _escribir_tabla(ws_dotacion, 1, COLUMNAS_DOTACION, dotacion)

    ws_proc = wb.create_sheet("Procedimientos")
    _escribir_tabla(ws_proc, 1, COLUMNAS_PROCEDIMIENTOS, procedimientos)

    return wb


def generar_reporte_excel(output_path: Path | None = None) -> Path:
    """Genera el Excel de Capacitacion y lo guarda en disco. No requiere el
    backend corriendo — consulta la base SQLite directamente. Devuelve la
    ruta del archivo generado."""
    datos = obtener_datos_reporte()
    if not datos["dotacion"]:
        raise RuntimeError(
            "Todavia no se cargo ninguna dotacion. Correr primero "
            "Capacitacion/normalizar_capacitacion.py."
        )

    wb = construir_workbook(datos["dotacion"], datos["nuevos"], datos["cambios"], datos["procedimientos"])

    if output_path is None:
        REPORTES_DIR.mkdir(parents=True, exist_ok=True)
        nombre = nombre_archivo_reporte(datos["mes_reporte"], datos["anio_reporte"])
        output_path = REPORTES_DIR / nombre
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
