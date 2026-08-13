import shutil
from datetime import datetime
from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.exc import OperationalError

from backend.auth.dependencies import get_current_user, require_roles
from backend.config import PROJECT_ROOT
from backend.database.models import User, UserRole
from backend.dashboards.capacitacion.db import (
    TABLE_CAMBIOS_CARGO,
    TABLE_DOTACION,
    TABLE_NUEVOS_INGRESOS,
    TABLE_PROCEDIMIENTOS,
    engine,
)
from backend.dashboards.capacitacion.normalizar import CLASIFICACION_DEFAULT, MESES_NOMBRE, procesar_archivos
from backend.dashboards.capacitacion.schemas import (
    ActualizacionOut,
    CambioCargoOut,
    CargoRevisarOut,
    DotacionOut,
    ProcedimientoIn,
    ProcedimientoOut,
)

UPLOADS_DIR = PROJECT_ROOT / "data" / "uploads" / "capacitacion"

router = APIRouter(prefix="/dashboards/capacitacion", tags=["dashboards:capacitacion"])

NO_DATA_ERROR = HTTPException(
    status_code=status.HTTP_404_NOT_FOUND,
    detail="Todavia no se cargaron dotaciones. Usa 'Actualizar datos' en el dashboard.",
)


def _query_con_clasificacion(table: str) -> list[dict]:
    """Trae todas las filas de `table` (dotacion/nuevos_ingresos/cambios_cargo)
    con la Clasificacion de Procedimientos calculada EN VIVO via JOIN contra
    la tabla maestra — asi que editar un codigo en 'Procedimientos' se ve
    reflejado al instante en todas las vistas y en el Excel exportado, sin
    tener que volver a subir las dotaciones."""
    sql = (
        f"SELECT d.*, COALESCE(p.Codigos, :default) AS Clasificacion "
        f"FROM {table} d LEFT JOIN {TABLE_PROCEDIMIENTOS} p ON d.Funcion = p.Funcion"
    )
    try:
        with engine.connect() as conn:
            rows = conn.execute(text(sql), {"default": CLASIFICACION_DEFAULT}).mappings().all()
    except OperationalError:
        raise NO_DATA_ERROR
    return [dict(row) for row in rows]


def _cargos_revisar_en_vivo() -> list[dict]:
    """Funciones de la dotacion actual que no existen en la tabla maestra —
    calculado en vivo, para que clasificar una desde la UI la saque de esta
    lista al instante."""
    sql = (
        f"SELECT d.Funcion AS Funcion, COUNT(*) AS Cantidad_Personas "
        f"FROM {TABLE_DOTACION} d LEFT JOIN {TABLE_PROCEDIMIENTOS} p ON d.Funcion = p.Funcion "
        f"WHERE p.Funcion IS NULL GROUP BY d.Funcion ORDER BY d.Funcion"
    )
    try:
        with engine.connect() as conn:
            rows = conn.execute(text(sql)).mappings().all()
    except OperationalError:
        raise NO_DATA_ERROR
    return [dict(row) for row in rows]


@router.get("/dotacion", response_model=list[DotacionOut])
def get_dotacion(_user: User = Depends(get_current_user)):
    return _query_con_clasificacion(TABLE_DOTACION)


@router.get("/nuevos-ingresos", response_model=list[DotacionOut])
def get_nuevos_ingresos(_user: User = Depends(get_current_user)):
    return _query_con_clasificacion(TABLE_NUEVOS_INGRESOS)


@router.get("/cambios-cargo", response_model=list[CambioCargoOut])
def get_cambios_cargo(_user: User = Depends(get_current_user)):
    return _query_con_clasificacion(TABLE_CAMBIOS_CARGO)


@router.get("/cargos-revisar", response_model=list[CargoRevisarOut])
def get_cargos_revisar(_user: User = Depends(get_current_user)):
    return _cargos_revisar_en_vivo()


@router.get("/procedimientos", response_model=list[ProcedimientoOut])
def get_procedimientos(_user: User = Depends(get_current_user)):
    with engine.connect() as conn:
        rows = conn.execute(text(f"SELECT * FROM {TABLE_PROCEDIMIENTOS} ORDER BY Funcion")).mappings().all()
    return [dict(row) for row in rows]


@router.put("/procedimientos", response_model=ProcedimientoOut)
def guardar_procedimiento(
    body: ProcedimientoIn,
    _admin: User = Depends(require_roles(UserRole.ADMINISTRADOR)),
):
    """Agrega o edita una fila de la tabla maestra Funcion -> Codigos. No pisa
    las otras tablas — un admin puede clasificar un cargo nuevo en cualquier
    momento, sin necesidad de volver a subir las dotaciones."""
    funcion = body.funcion.strip()
    codigos = body.codigos.strip()
    if not funcion or not codigos:
        raise HTTPException(status_code=400, detail="Funcion y Codigos son obligatorios")

    with engine.begin() as conn:
        conn.execute(
            text(f"INSERT OR REPLACE INTO {TABLE_PROCEDIMIENTOS} (Funcion, Codigos) VALUES (:f, :c)"),
            {"f": funcion, "c": codigos},
        )
    return {"Funcion": funcion, "Codigos": codigos}


@router.post("/actualizar", response_model=ActualizacionOut)
async def actualizar_datos(
    archivo_actual: UploadFile = File(...),
    archivo_anterior: UploadFile = File(...),
    mes: int = Form(...),
    anio: int = Form(...),
    _admin: User = Depends(require_roles(UserRole.ADMINISTRADOR)),
):
    """Sube las 2 dotaciones (mes del reporte + mes anterior), las procesa y
    reemplaza las 3 tablas calculadas — respaldando antes la base anterior.
    Solo admin, porque pisa los datos que ven todos los usuarios."""
    for archivo, etiqueta in ((archivo_actual, "actual"), (archivo_anterior, "anterior")):
        if not archivo.filename or not archivo.filename.lower().endswith(".xlsx"):
            raise HTTPException(
                status_code=400,
                detail=f"El archivo de dotacion {etiqueta} debe ser un Excel (.xlsx)",
            )
    if not (1 <= mes <= 12):
        raise HTTPException(status_code=400, detail="El mes debe estar entre 1 y 12")

    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    destino_actual = UPLOADS_DIR / f"{timestamp}_actual_{archivo_actual.filename}"
    destino_anterior = UPLOADS_DIR / f"{timestamp}_anterior_{archivo_anterior.filename}"
    with destino_actual.open("wb") as f:
        shutil.copyfileobj(archivo_actual.file, f)
    with destino_anterior.open("wb") as f:
        shutil.copyfileobj(archivo_anterior.file, f)

    try:
        resultado = procesar_archivos(destino_actual, destino_anterior, mes, anio)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=(
                "No se pudieron procesar los archivos — revisa que ambos tengan una hoja "
                f"'Detalle...' con el formato esperado. Detalle: {exc}"
            ),
        )

    resultado["cargos_revisar"] = len(_cargos_revisar_en_vivo())
    return resultado


# ------------------------------------------------------------------------
# Exportar Excel — misma estructura de 3 hojas que
# "Capacitacion/Archivos Ejemplo/Reporte para capacitaciones - *.xlsx"
# ------------------------------------------------------------------------
COLUMNAS_DOTACION = [
    ("Cod_Personal", "N° pers."), ("Nombre_Completo", "Nombre completo"),
    ("Area_Personal", "Área de personal"), ("Sociedad", "Sociedad"),
    ("Unidad_Organizativa", "Unidad organizativa"), ("Funcion", "Función"),
    ("Gerencia", "Gerencia"), ("Subgerencia", "Subgerencia"),
    ("Fecha_Ingreso", "Fecha Ingreso"), ("Nombre_Superior", "Nombre del superior (GO)"),
    ("Nacionalidad", "Nacionalidad"), ("Mail", "Mail"),
    ("Clasificacion", "Clasificación de Procedimientos"),
]
COLUMNAS_CAMBIOS_CARGO = [
    ("Cod_Personal", "N° pers."), ("Nombre_Completo", "Nombre completo"),
    ("Area_Personal", "Área de personal"), ("Sociedad", "Sociedad"),
    ("Unidad_Organizativa", "Unidad organizativa"), ("Cargo_Anterior", "Cargo Anterior"),
    ("Funcion", "Función"), ("Gerencia", "Gerencia"), ("Subgerencia", "Subgerencia"),
    ("Fecha_Ingreso", "Fecha Ingreso"), ("Nombre_Superior", "Nombre del superior (GO)"),
    ("Nacionalidad", "Nacionalidad"), ("Mail", "Mail"),
    ("Clasificacion", "Clasificación de Procedimientos"),
]
COLUMNAS_PROCEDIMIENTOS = [("Funcion", "Función"), ("Codigos", "Procedimientos asignados")]


def _valor_celda(campo: str, valor):
    if campo == "Fecha_Ingreso" and isinstance(valor, str) and valor:
        try:
            return datetime.strptime(valor[:10], "%Y-%m-%d").date()
        except ValueError:
            return valor
    return valor


def _escribir_tabla(ws, start_row: int, columnas, filas: list[dict]) -> int:
    for j, (_, header) in enumerate(columnas, start=1):
        ws.cell(row=start_row, column=j, value=header)
    r = start_row + 1
    for fila in filas:
        for j, (campo, _) in enumerate(columnas, start=1):
            ws.cell(row=r, column=j, value=_valor_celda(campo, fila.get(campo)))
        r += 1
    return r


def _construir_workbook(dotacion, nuevos, cambios, procedimientos):
    wb = openpyxl.Workbook()

    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    fila = 2
    ws_resumen.cell(row=fila, column=1, value="Nuevos ingresos ")
    fila = _escribir_tabla(ws_resumen, fila + 1, COLUMNAS_DOTACION, nuevos)
    fila += 2
    ws_resumen.cell(row=fila, column=1, value="Cambios de cargo")
    _escribir_tabla(ws_resumen, fila + 1, COLUMNAS_CAMBIOS_CARGO, cambios)

    ws_dotacion = wb.create_sheet("Dotación CHTA")
    _escribir_tabla(ws_dotacion, 1, COLUMNAS_DOTACION, dotacion)

    ws_proc = wb.create_sheet("Procedimientos")
    _escribir_tabla(ws_proc, 1, COLUMNAS_PROCEDIMIENTOS, procedimientos)

    return wb


@router.get("/exportar-excel")
def exportar_excel(_user: User = Depends(get_current_user)):
    dotacion = _query_con_clasificacion(TABLE_DOTACION)
    if not dotacion:
        raise NO_DATA_ERROR
    nuevos = _query_con_clasificacion(TABLE_NUEVOS_INGRESOS)
    cambios = _query_con_clasificacion(TABLE_CAMBIOS_CARGO)
    with engine.connect() as conn:
        procedimientos = [
            dict(row) for row in conn.execute(
                text(f"SELECT * FROM {TABLE_PROCEDIMIENTOS} ORDER BY Funcion")
            ).mappings().all()
        ]

    mes_reporte = dotacion[0].get("Mes_Reporte")
    anio_reporte = dotacion[0].get("Anio_Reporte")
    mes_nombre = MESES_NOMBRE.get(mes_reporte, "")

    wb = _construir_workbook(dotacion, nuevos, cambios, procedimientos)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = f"Reporte para capacitaciones - {mes_nombre} {anio_reporte}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'},
    )
