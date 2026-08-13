"""
================================================================================
NORMALIZADOR DE SOBRETIEMPO PARA POWER BI
================================================================================
Convierte el Excel en bruto "Control de Sobretiempo" (hojas DETALLE 2,0 y
PPTO 2026) en un archivo Excel normalizado con 3 tablas listas para cargar
a Power BI:

    - Detalle        : transaccional, un registro por persona/concepto/mes
    - Presupuesto     : presupuesto 2026 despivoteado (mes en filas, no en columnas)
    - Resumen         : cruce Real vs Presupuesto por Ceco + Cuenta + Mes,
                         con variación, % de ejecución y estado semaforo

USO CADA MES:
    python normalizar_sobretiempo.py "ruta/al/nuevo_archivo.xlsx" "ruta/salida.xlsx"

Si no se indican rutas, usa los valores por defecto definidos en INPUT_PATH /
OUTPUT_PATH más abajo.
================================================================================
"""

import sys
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# ------------------------------------------------------------------------
# CONFIGURACIÓN
# ------------------------------------------------------------------------
INPUT_PATH = "/mnt/user-data/uploads/Control_de_Sobretiempo_2026.xlsx"
OUTPUT_PATH = "/mnt/user-data/outputs/Sobretiempo_Normalizado.xlsx"

SHEET_DETALLE = "DETALLE 2,0"
SHEET_PPTO = "PPTO 2026"

MESES_ORDEN = {
    "Enero": 1, "Febrero": 2, "Marzo": 3, "Abril": 4, "Mayo": 5, "Junio": 6,
    "Julio": 7, "Agosto": 8, "Septiembre": 9, "Octubre": 10,
    "Noviembre": 11, "Diciembre": 12,
}
MESES_INV = {v: k for k, v in MESES_ORDEN.items()}
# Etiqueta de mes con número al inicio (ej. "01-Enero") para que las tablas
# ordenen bien como texto en Power BI sin tener que configurar "Ordenar por
# columna" sobre Mes_Nombre
MESES_LABEL = {v: f"{v:02d}-{k}" for k, v in MESES_ORDEN.items()}

# Códigos de Sociedad válidos (la hoja PPTO 2026 los usa en su tabla de
# detalle; sirven además para descartar de forma robusta la tabla-resumen
# que viene más abajo en la misma hoja, que reutiliza las mismas columnas
# con otro contenido)
SOCIEDAD_MAP = {
    "CC": "Chilquinta Energía S.A.",
    "CE": "Chilquinta Distribución S",
    "CS": "Chilquinta Servicios S.A.",
    "CX": "Chilquinta Transmisión S.",
    "CA": "Energ. de Casablanca S.A.",
    "LT": "C. Eléc. del Litoral S.A.",
    "LN": "Luzlinares S.A.",
    "LP": "Luzparral S.A.",
}


# ------------------------------------------------------------------------
# 1. LECTURA Y LIMPIEZA DE "DETALLE 2,0"
# ------------------------------------------------------------------------
def cargar_detalle(path):
    """
    La hoja trae filas basura arriba (filas 2-36, una matriz de referencia
    con Importe=0 que no es data real) y filas vacías al final. Solo son
    datos reales las filas donde 'Cód SAP' es numérico.
    """
    raw = pd.read_excel(path, sheet_name=SHEET_DETALLE, header=0)

    # Filtrar solo filas con Cód SAP numérico (descarta filas basura / vacías)
    raw = raw[pd.to_numeric(raw["Cód SAP"], errors="coerce").notna()].copy()

    # La hoja trae "Gerencia"/"Subgerencia"/"Unidad organizativa" DOS veces
    # (columnas duplicadas). Las segundas (posiciones 19-21) son la
    # jerarquía organizacional "oficial", la misma que usa PPTO 2026 para
    # cruzar por Ceco+Cuenta. Se usan esas.
    cols = list(raw.columns)
    gerencia_col = cols[19]
    subgerencia_col = cols[20]
    unidad_col = cols[21]

    df = pd.DataFrame({
        "Cod_SAP": raw["Cód SAP"].astype("Int64"),
        "Nombre_Personal": raw["Nombre de personal"],
        "RUT": raw["RUT"],
        "Sociedad": raw["Sociedad"],
        "Division_Personal": raw["División de personal"],
        "Gerencia": raw[gerencia_col],
        "Subgerencia": raw[subgerencia_col],
        "Unidad_Organizativa": raw[unidad_col],
        "Cargo": raw["Cargo"],
        "Area_Personal": raw["Área de Personal"],
        "Ceco": raw["Ceco"].fillna("").astype(str).str.strip(),
        "Cuenta_Contable": pd.to_numeric(raw["Cuenta contable"], errors="coerce").astype("Int64"),
        "OI": raw["OI"],
        "PEP": raw["PEP"],
        "Codigo_Concepto": raw["CC-n."],
        "Concepto": raw["Texto expl.CC-nómina"],
        "Clasif_Haber": raw["Clasif Haber"],
        "Clasificacion": raw["Clasificación"],
        "Cantidad_Horas": pd.to_numeric(raw["Cantid."], errors="coerce"),
        "Importe": pd.to_numeric(raw["   Importe"], errors="coerce"),
        "Validacion": raw["Validación"],
        "Anio": pd.to_numeric(raw[" Año"], errors="coerce").astype("Int64"),
        "Mes_Nombre": raw["Mes"],
    })

    df["Mes_Num"] = df["Mes_Nombre"].map(MESES_ORDEN).astype("Int64")
    df["Mes_Orden"] = df["Mes_Num"].map(MESES_LABEL)
    df["Fecha"] = df.apply(
        lambda r: date(int(r["Anio"]), int(r["Mes_Num"]), 1)
        if pd.notna(r["Anio"]) and pd.notna(r["Mes_Num"]) else None, axis=1
    )

    # Subcuenta = últimos 3 dígitos de la cuenta contable (mismo criterio que
    # la columna "Scta" de PPTO 2026)
    df["Subcuenta"] = df["Cuenta_Contable"].apply(
        lambda c: str(int(c))[-3:] if pd.notna(c) else None
    )

    # Ceco vacío = sobretiempo cargado directo a un PEP/proyecto en vez de
    # a un centro de costo (no tiene presupuesto CECO asociado)
    df["Ceco"] = df["Ceco"].replace("", "SIN CECO (Cargo a Proyecto)")

    return df


# ------------------------------------------------------------------------
# 4. ENRIQUECER DETALLE: flag de presupuesto asignado + mes cerrado
#    (mismos criterios que la hoja Resumen, para que ambas hojas se lean
#    de forma consistente en Power BI sin necesidad de medidas DAX)
# ------------------------------------------------------------------------
def enriquecer_detalle(df_detalle, df_ppto):
    # Mismo criterio validado contra la presentación oficial: "con
    # presupuesto asignado" se define a nivel Ceco + Cuenta Contable.
    tiene_ppto = (
        df_ppto.groupby(["Ceco", "Cuenta_Contable"])["Presupuesto"].sum() > 0
    ).rename("Con_Presupuesto_Asignado").reset_index()
    df = df_detalle.merge(tiene_ppto, on=["Ceco", "Cuenta_Contable"], how="left")
    df["Con_Presupuesto_Asignado"] = df["Con_Presupuesto_Asignado"].fillna(False)

    ultimo_mes_cerrado = int(df.loc[df["Importe"] > 0, "Mes_Num"].max())
    df["Mes_Cerrado"] = df["Mes_Num"] <= ultimo_mes_cerrado
    return df


# ------------------------------------------------------------------------
# 4b. RESUMEN EJECUTIVO POR GERENCIA + SUBGERENCIA
#     (replica la tabla de seguimiento que usa Gerencia de Personas: ver
#     Control_Sobretiempo_presentación.pptx, diapositiva 3-4)
#
#     Metodología validada contra la presentación oficial:
#       - Solo se consideran cuentas CON presupuesto asignado (el gasto
#         cargado a PEP/proyecto, sin Ceco, queda fuera del avance
#         presupuestario -aunque sigue disponible en la hoja Detalle-)
#       - El "Presupuesto aprobado" es un monto FIJO ANUAL por Gerencia +
#         Subgerencia (no se compara contra un acumulado de presupuesto
#         mes a mes, sino contra el total aprobado para el año)
#       - "Trabajadores con HE" = cantidad de personas (Cód SAP) distintas
#         con sobretiempo, mensual y acumulado en el año
# ------------------------------------------------------------------------
KEY_GS = ["Gerencia", "Subgerencia"]


def construir_resumen_gerencia(df_detalle, df_ppto):
    # Nota metodológica importante: el filtro "con presupuesto asignado" se
    # evalúa a nivel Ceco + Cuenta Contable (columna Con_Presupuesto_Asignado
    # de Detalle). Se validó contra la presentación oficial que así es como
    # Gerencia de Personas define "cuentas presupuestadas": si el Ceco+Cuenta
    # tiene presupuesto en cualquier unidad, todo su gasto real cuenta
    # -aunque la unidad puntual que gastó no tenga su propia línea de
    # presupuesto-.
    det = df_detalle[df_detalle["Con_Presupuesto_Asignado"]].copy()

    anio = int(df_detalle["Anio"].dropna().iloc[0]) if df_detalle["Anio"].notna().any() else 2026
    ultimo_mes_cerrado = int(df_detalle.loc[df_detalle["Importe"] > 0, "Mes_Num"].max())

    ppto_anual = (
        df_ppto.groupby(KEY_GS)["Presupuesto"].sum()
        .rename("Presupuesto_Total_Anual").reset_index()
    )
    real_mes = (
        det.groupby(KEY_GS + ["Mes_Num"])["Importe"].sum()
        .rename("Importe_Real_Mes").reset_index()
    )
    head_mes = (
        det.groupby(KEY_GS + ["Mes_Num"])["Cod_SAP"].nunique()
        .rename("Trabajadores_Con_HE_Mes").reset_index()
    )

    # Universo: todas las combinaciones Sociedad+Gerencia+Subgerencia (de
    # cualquiera de las 2 fuentes) x los 12 meses del año
    combos = pd.concat([ppto_anual[KEY_GS], real_mes[KEY_GS]]).drop_duplicates()
    meses = pd.DataFrame({"Mes_Num": list(range(1, 13))})
    combos["_k"] = 1
    meses["_k"] = 1
    universo = combos.merge(meses, on="_k").drop(columns="_k")
    universo["Anio"] = anio

    r = (
        universo
        .merge(ppto_anual, on=KEY_GS, how="left")
        .merge(real_mes, on=KEY_GS + ["Mes_Num"], how="left")
        .merge(head_mes, on=KEY_GS + ["Mes_Num"], how="left")
    )
    r["Presupuesto_Total_Anual"] = r["Presupuesto_Total_Anual"].fillna(0)
    r["Importe_Real_Mes"] = r["Importe_Real_Mes"].fillna(0)
    r["Trabajadores_Con_HE_Mes"] = r["Trabajadores_Con_HE_Mes"].fillna(0).astype(int)

    r["Mes_Nombre"] = r["Mes_Num"].map(MESES_INV)
    r["Mes_Orden"] = r["Mes_Num"].map(MESES_LABEL)
    r["Fecha"] = r.apply(lambda x: date(int(x["Anio"]), int(x["Mes_Num"]), 1), axis=1)
    r["Mes_Cerrado"] = r["Mes_Num"] <= ultimo_mes_cerrado
    r = r.sort_values(KEY_GS + ["Mes_Num"]).reset_index(drop=True)

    # --- Acumulado de gasto real (YTD) ---
    r["Real_Acumulado"] = r.groupby(KEY_GS, sort=False)["Importe_Real_Mes"].cumsum()

    # --- Headcount acumulado (personas distintas con HE en lo que va del
    # año -> NO es la suma de los headcounts mensuales, porque la misma
    # persona puede repetirse mes a mes; se cuenta 1 vez desde su primer
    # mes con sobretiempo) ---
    primera_aparicion = (
        det.groupby(KEY_GS + ["Cod_SAP"])["Mes_Num"].min()
        .rename("Primer_Mes").reset_index()
    )
    altas_por_mes = (
        primera_aparicion.groupby(KEY_GS + ["Primer_Mes"]).size()
        .rename("Altas").reset_index()
        .rename(columns={"Primer_Mes": "Mes_Num"})
    )
    r = r.merge(altas_por_mes, on=KEY_GS + ["Mes_Num"], how="left")
    r["Altas"] = r["Altas"].fillna(0).astype(int)
    r["Trabajadores_Con_HE_Acumulado"] = r.groupby(KEY_GS, sort=False)["Altas"].cumsum()
    r = r.drop(columns="Altas")

    # --- Saldo, % ocupado y estado (contra el presupuesto ANUAL fijo) ---
    r["Saldo_Disponible"] = r["Presupuesto_Total_Anual"] - r["Real_Acumulado"]
    r["Pct_Ocupado"] = np.where(
        r["Presupuesto_Total_Anual"] > 0, r["Real_Acumulado"] / r["Presupuesto_Total_Anual"], np.nan
    )

    def estado(saldo, ppto, mes_cerrado):
        if ppto == 0:
            return "Sin Presupuesto"
        if not mes_cerrado:
            return "Mes Futuro (sin cierre)"
        return "Saldo a Favor" if saldo >= 0 else "Saldo en Contra"

    r["Estado"] = [estado(s, p, m) for s, p, m in
                   zip(r["Saldo_Disponible"], r["Presupuesto_Total_Anual"], r["Mes_Cerrado"])]

    cols_order = KEY_GS + [
        "Anio", "Mes_Num", "Mes_Nombre", "Mes_Orden", "Fecha", "Mes_Cerrado",
        "Presupuesto_Total_Anual", "Importe_Real_Mes", "Real_Acumulado", "Saldo_Disponible",
        "Pct_Ocupado", "Trabajadores_Con_HE_Mes", "Trabajadores_Con_HE_Acumulado", "Estado",
    ]
    return r[cols_order].sort_values(KEY_GS + ["Mes_Num"]).reset_index(drop=True)


# ------------------------------------------------------------------------
# 2. LECTURA Y DESPIVOTEO DE "PPTO 2026"
# ------------------------------------------------------------------------
def cargar_presupuesto(path):
    """
    La hoja PPTO 2026 trae el detalle real de presupuesto en la parte
    superior (formato ancho: un mes por columna). Más abajo en la MISMA
    hoja hay una segunda tabla-resumen (por Sociedad/Gerencia) que reutiliza
    los mismos nombres de columna con otro contenido. Para descartarla de
    forma robusta (sin depender de un número de fila fijo que puede
    correrse mes a mes) nos quedamos solo con filas cuya 'Sociedad' sea uno
    de los códigos válidos de 2 letras (CA/CC/CE/CS/CX/LN/LP/LT): ese
    patrón solo aparece en la tabla de detalle real.
    """
    raw = pd.read_excel(path, sheet_name=SHEET_PPTO, header=0)
    raw = raw[raw["Sociedad"].isin(SOCIEDAD_MAP.keys())].copy()

    meses_cols = [f"Suma de {m}" for m in MESES_ORDEN]
    meses_cols = [c for c in meses_cols if c in raw.columns]

    id_cols = {
        "Centro de Costo": "Ceco",
        "Cuenta": "Cuenta_Contable",
        "Scta": "Subcuenta",
        "Sociedad": "Sociedad",
        "Gerencia": "Gerencia",
        "Subgerencia": "Subgerencia",
        "Unidad organizativa": "Unidad_Organizativa",
    }
    base = raw[list(id_cols.keys()) + meses_cols].rename(columns=id_cols)
    base["Ceco"] = base["Ceco"].astype(str).str.strip()
    base["Cuenta_Contable"] = pd.to_numeric(base["Cuenta_Contable"], errors="coerce").round().astype("Int64")
    base["Subcuenta"] = pd.to_numeric(base["Subcuenta"], errors="coerce").round().astype("Int64").astype(str).str.zfill(3)
    # Estandarizar Sociedad: mismo nombre completo que usa la hoja Detalle
    base["Sociedad"] = base["Sociedad"].map(SOCIEDAD_MAP)

    # Despivotear: de "un mes por columna" a "una fila por mes"
    long_df = base.melt(
        id_vars=[c for c in id_cols.values()],
        value_vars=meses_cols,
        var_name="Mes_Nombre",
        value_name="Presupuesto",
    )
    long_df["Mes_Nombre"] = long_df["Mes_Nombre"].str.replace("Suma de ", "", regex=False)
    long_df["Mes_Num"] = long_df["Mes_Nombre"].map(MESES_ORDEN).astype("Int64")
    long_df["Mes_Orden"] = long_df["Mes_Num"].map(MESES_LABEL)
    long_df["Presupuesto"] = pd.to_numeric(long_df["Presupuesto"], errors="coerce").fillna(0)
    long_df["Anio"] = 2026

    # Como un mismo Ceco+Cuenta puede repartirse en varias filas por distinta
    # Gerencia/Subgerencia/Unidad, se agrega a nivel Ceco+Cuenta+Mes (que es
    # la llave de cruce pedida) y también se conserva el detalle por unidad.
    long_df["Fecha"] = long_df.apply(
        lambda r: date(int(r["Anio"]), int(r["Mes_Num"]), 1) if pd.notna(r["Mes_Num"]) else None,
        axis=1,
    )

    cols_order = ["Ceco", "Cuenta_Contable", "Subcuenta", "Sociedad", "Gerencia",
                  "Subgerencia", "Unidad_Organizativa", "Anio", "Mes_Num",
                  "Mes_Nombre", "Mes_Orden", "Fecha", "Presupuesto"]
    return long_df[cols_order].sort_values(["Ceco", "Cuenta_Contable", "Mes_Num"]).reset_index(drop=True)


# ------------------------------------------------------------------------
# 3. RESUMEN: CRUCE REAL VS PRESUPUESTO
#    Llave: Sociedad + Ceco + Cuenta + Gerencia + Subgerencia + Unidad + Mes
#    (Ceco+Cuenta es la llave pedida para cruzar; se agrega Gerencia /
#    Subgerencia / Unidad porque en los datos un mismo Ceco reparte
#    presupuesto entre varias Gerencias -> sin este detalle no se puede
#    saber qué área es "responsable" de cada línea de presupuesto)
#
#    Trae TODO precalculado para que en Power BI no haga falta escribir
#    ninguna medida DAX: monto del mes, acumulado del año (YTD), saldo
#    mensual y acumulado, % de ejecución mensual y acumulado, estado y
#    flag de mes cerrado.
# ------------------------------------------------------------------------
KEY_ORG = ["Sociedad", "Ceco", "Cuenta_Contable", "Subcuenta",
           "Gerencia", "Subgerencia", "Unidad_Organizativa"]


def construir_resumen(df_detalle, df_ppto):
    real_agg = (
        df_detalle.groupby(KEY_ORG + ["Anio", "Mes_Num"], dropna=False)
        .agg(Importe_Real=("Importe", "sum"), Horas_Real=("Cantidad_Horas", "sum"))
        .reset_index()
    )
    ppto_agg = (
        df_ppto.groupby(KEY_ORG + ["Anio", "Mes_Num"], dropna=False)
        .agg(Presupuesto=("Presupuesto", "sum"))
        .reset_index()
    )

    # Universo completo: todas las combinaciones que existen en cualquiera
    # de las dos fuentes x los 12 meses del año -> permite ver meses
    # futuros ya presupuestados aunque aún no haya gasto real, y gasto
    # real sin presupuesto asignado ("Sin Presupuesto")
    combos = pd.concat([real_agg[KEY_ORG], ppto_agg[KEY_ORG]]).drop_duplicates()

    anio = int(df_detalle["Anio"].dropna().iloc[0]) if df_detalle["Anio"].notna().any() else 2026
    meses = pd.DataFrame({"Mes_Num": list(range(1, 13))})
    combos["_k"] = 1
    meses["_k"] = 1
    universo = combos.merge(meses, on="_k").drop(columns="_k")
    universo["Anio"] = anio

    resumen = (
        universo
        .merge(real_agg, on=KEY_ORG + ["Anio", "Mes_Num"], how="left")
        .merge(ppto_agg, on=KEY_ORG + ["Anio", "Mes_Num"], how="left")
    )
    resumen["Importe_Real"] = resumen["Importe_Real"].fillna(0)
    resumen["Horas_Real"] = resumen["Horas_Real"].fillna(0)
    resumen["Presupuesto"] = resumen["Presupuesto"].fillna(0)  # Sin Presupuesto -> 0

    # ¿Esta línea (Sociedad+Ceco+Cuenta+Gerencia+Subgerencia+Unidad) tiene
    # presupuesto asignado en algún mes del año?
    tiene_ppto = ppto_agg.groupby(KEY_ORG)["Presupuesto"].sum()
    tiene_ppto = (tiene_ppto > 0).rename("Con_Presupuesto_Asignado")
    resumen = resumen.merge(tiene_ppto, on=KEY_ORG, how="left")
    resumen["Con_Presupuesto_Asignado"] = resumen["Con_Presupuesto_Asignado"].fillna(False)

    resumen["Mes_Nombre"] = resumen["Mes_Num"].map(MESES_INV)
    resumen["Mes_Orden"] = resumen["Mes_Num"].map(MESES_LABEL)
    resumen["Fecha"] = resumen.apply(lambda r: date(int(r["Anio"]), int(r["Mes_Num"]), 1), axis=1)

    # Ordenar por grupo + mes ANTES de calcular acumulados
    resumen = resumen.sort_values(KEY_ORG + ["Mes_Num"]).reset_index(drop=True)

    # --- Acumulados YTD (precalculados, sin necesidad de DAX time-intelligence) ---
    grp = resumen.groupby(KEY_ORG, sort=False)
    resumen["Presupuesto_Acumulado"] = grp["Presupuesto"].cumsum()
    resumen["Real_Acumulado"] = grp["Importe_Real"].cumsum()
    resumen["Horas_Real_Acumulado"] = grp["Horas_Real"].cumsum()

    # --- Saldo (a favor si es positivo, en contra si es negativo) ---
    resumen["Saldo_Mes"] = resumen["Presupuesto"] - resumen["Importe_Real"]
    resumen["Saldo_Acumulado"] = resumen["Presupuesto_Acumulado"] - resumen["Real_Acumulado"]

    # --- % de ejecución mensual y acumulado ---
    resumen["Pct_Ejecucion_Mes"] = np.where(
        resumen["Presupuesto"] > 0, resumen["Importe_Real"] / resumen["Presupuesto"], np.nan
    )
    resumen["Pct_Ejecucion_Acumulado"] = np.where(
        resumen["Presupuesto_Acumulado"] > 0,
        resumen["Real_Acumulado"] / resumen["Presupuesto_Acumulado"], np.nan
    )

    # --- Último mes con datos reales cargados (para distinguir "aún no
    # llega la data de ese mes" de "gastó $0 ese mes") ---
    ultimo_mes_cerrado = int(df_detalle.loc[df_detalle["Importe"] > 0, "Mes_Num"].max())
    resumen["Mes_Cerrado"] = resumen["Mes_Num"] <= ultimo_mes_cerrado

    def estado(saldo, con_ppto, ppto_mes, mes_cerrado):
        if not con_ppto:
            return "Sin Presupuesto"
        if not mes_cerrado:
            return "Mes Futuro (sin cierre)"
        if ppto_mes == 0:
            return "Sin Presupuesto (mes)"
        return "Saldo a Favor" if saldo >= 0 else "Saldo en Contra"

    resumen["Estado_Mes"] = [
        estado(s, c, p, m) for s, c, p, m in zip(
            resumen["Saldo_Mes"], resumen["Con_Presupuesto_Asignado"],
            resumen["Presupuesto"], resumen["Mes_Cerrado"])
    ]
    resumen["Estado_Acumulado"] = [
        estado(s, c, p, m) for s, c, p, m in zip(
            resumen["Saldo_Acumulado"], resumen["Con_Presupuesto_Asignado"],
            resumen["Presupuesto_Acumulado"], resumen["Mes_Cerrado"])
    ]

    cols_order = KEY_ORG + [
        "Anio", "Mes_Num", "Mes_Nombre", "Mes_Orden", "Fecha", "Mes_Cerrado",
        "Importe_Real", "Horas_Real", "Presupuesto", "Saldo_Mes", "Pct_Ejecucion_Mes", "Estado_Mes",
        "Real_Acumulado", "Horas_Real_Acumulado", "Presupuesto_Acumulado", "Saldo_Acumulado",
        "Pct_Ejecucion_Acumulado", "Estado_Acumulado", "Con_Presupuesto_Asignado",
    ]
    return resumen[cols_order].sort_values(KEY_ORG + ["Mes_Num"]).reset_index(drop=True)


# ------------------------------------------------------------------------
# 5. ESCRITURA DEL EXCEL DE SALIDA CON FORMATO
# ------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def formatear_hoja(ws, df):
    ws.freeze_panes = "A2"
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=j)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).values[:200]])
        ws.column_dimensions[get_column_letter(j)].width = min(max(max_len + 2, 10), 45)

        # Formato numérico para columnas de plata / porcentaje
        col_lower = col.lower()
        for i in range(2, len(df) + 2):
            c = ws.cell(row=i, column=j)
            c.font = BODY_FONT
            c.border = BORDER
            if col in ("Importe", "Importe_Real", "Presupuesto", "Saldo_Mes", "Saldo_Acumulado",
                       "Real_Acumulado", "Presupuesto_Acumulado", "Presupuesto_Total_Anual",
                       "Importe_Real_Mes", "Saldo_Disponible"):
                c.number_format = "#,##0"
            elif col in ("Pct_Ejecucion_Mes", "Pct_Ejecucion_Acumulado", "Pct_Ocupado"):
                c.number_format = "0.0%"
            elif col in ("Fecha",):
                c.number_format = "yyyy-mm-dd"


def guardar_excel(df_detalle, df_ppto, df_resumen, df_resumen_ger, output_path):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_detalle.to_excel(writer, sheet_name="Detalle", index=False)
        df_ppto.to_excel(writer, sheet_name="Presupuesto", index=False)
        df_resumen.to_excel(writer, sheet_name="Resumen", index=False)
        df_resumen_ger.to_excel(writer, sheet_name="Resumen_Gerencia", index=False)

    wb = openpyxl.load_workbook(output_path)
    formatear_hoja(wb["Detalle"], df_detalle)
    formatear_hoja(wb["Presupuesto"], df_ppto)
    formatear_hoja(wb["Resumen"], df_resumen)
    formatear_hoja(wb["Resumen_Gerencia"], df_resumen_ger)
    wb.save(output_path)


# ------------------------------------------------------------------------
# MAIN
# ------------------------------------------------------------------------
def main(input_path=INPUT_PATH, output_path=OUTPUT_PATH):
    print(f"Leyendo: {input_path}")
    df_detalle = cargar_detalle(input_path)
    df_ppto = cargar_presupuesto(input_path)
    df_detalle = enriquecer_detalle(df_detalle, df_ppto)
    df_resumen = construir_resumen(df_detalle, df_ppto)
    df_resumen_ger = construir_resumen_gerencia(df_detalle, df_ppto)

    print(f"  Detalle:            {len(df_detalle):,} filas")
    print(f"  Presupuesto:        {len(df_ppto):,} filas")
    print(f"  Resumen:            {len(df_resumen):,} filas")
    print(f"  Resumen_Gerencia:   {len(df_resumen_ger):,} filas")

    guardar_excel(df_detalle, df_ppto, df_resumen, df_resumen_ger, output_path)
    print(f"Archivo generado: {output_path}")


if __name__ == "__main__":
    inp = sys.argv[1] if len(sys.argv) > 1 else INPUT_PATH
    out = sys.argv[2] if len(sys.argv) > 2 else OUTPUT_PATH
    main(inp, out)
